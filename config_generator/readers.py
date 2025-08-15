# ==============================================================================
# TableCompiler
# Copyright (c) 2025, Alex Liao. All rights reserved.
#
# This file is part of the TableCompiler project, a tool designed to
# compile Excel configuration sheets into type-safe code and binary data
# for high-performance projects.
# ==============================================================================

# /config_generator/readers.py
import os
import json
import openpyxl
import re
from .models import ConfigTable, ConfigRow

# 用于解析 list(int) 或 array(Item) 这样的语法
TYPE_STRING_REGEX = re.compile(r"^(list|set|array)\((.*)\)$")

def parse_type_string(type_str):
    """
    解析集合类型字符串。
    
    Args:
        type_str: 待解析的字符串, 如 'list(Item)'。
    
    Returns:
        一个元组 (collection_type, inner_type_string)，例如 ('list', 'Item')。
        如果不是集合类型，则返回 (None, original_string)。
    """
    if not isinstance(type_str, str):
        return None, type_str
    match = TYPE_STRING_REGEX.match(type_str.strip())
    if match:
        return match.group(1), match.group(2)
    return None, type_str

class TypeSystem:
    """
    管理所有已加载的类型定义和默认解析模式。
    它会递归加载所有导入的 .innertypesdef.json 文件，并提供一个统一的接口来查询类型。
    """
    def __init__(self, metadata_dir: str):
        self._loaded_types = {}
        self._default_schemas = {}
        self.metadata_dir = metadata_dir # 存储元数据根目录

    def load_type_def(self, rel_path_from_meta: str, silent: bool = False):
        """
        加载并解析一个 .innertypesdef.json 文件。
        路径总是相对于 metadata 根目录，且不包含后缀。
        """
        full_path = os.path.join(self.metadata_dir, f"{rel_path_from_meta}.innertypesdef.json")
        abs_path = os.path.abspath(full_path)

        if abs_path in self._loaded_types.get("@@files", set()):
            return
        
        if not silent:
            print(f"  -> 加载类型定义: {os.path.basename(full_path)}")
        
        try:
            with open(abs_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到导入的类型定义文件: {full_path}")
        except json.JSONDecodeError as e:
            raise ValueError(f"'{abs_path}' 文件错误: {e}")
        
        self._loaded_types.setdefault("@@files", set()).add(abs_path)

        # 递归加载导入的类型，路径同样相对于 metadata 根目录
        for p in data.get("ImportTypes", []):
            self.load_type_def(p, silent)
        
        for name, info in data.get("TypeDefines", {}).items():
            self._loaded_types[name] = info
        
        for type_expr, schema in data.get("SourceSchemas", {}).items():
            self._default_schemas[type_expr] = schema

    def get_type(self, name: str) -> dict:
        """根据名称获取类型定义。"""
        if name in self._loaded_types:
            return self._loaded_types[name]
        
        coll, _ = parse_type_string(name)
        if coll or name in ["int", "long", "string", "bool", "float"]:
            return {"TargetType": name}
            
        raise ValueError(f"类型 '{name}' 未在任何 .typedef 或 .innertypesdef 文件中定义。")
    
    def get_default_schema(self, type_expr: str) -> dict | None:
        """根据完整的类型表达式获取默认的解析模式。"""
        return self._default_schemas.get(type_expr)

    def get_all_custom_type_names(self) -> list[str]:
        """获取所有已加载的自定义类型名称。"""
        return [k for k in self._loaded_types.keys() if k != "@@files"]

class ConfigReader:
    """
    读取所有配置并将其解析为语言无关的中间数据结构 (ConfigTable)。
    """
    def __init__(self, input_dir: str, metadata_dir: str, typedef_suffix: str):
        self.input_dir = input_dir
        self.metadata_dir = metadata_dir
        self.typedef_suffix = typedef_suffix
        self.type_system = TypeSystem(self.metadata_dir)
        self._scan_all_innertypes()

    def _scan_all_innertypes(self):
        """预扫描并加载 metadata 目录下的所有 innertype 定义。"""
        suffix = ".innertypesdef.json"
        for root, _, files in os.walk(self.metadata_dir):
            for file in files:
                if file.endswith(suffix):
                    try:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(full_path, self.metadata_dir)
                        
                        # 可靠地移除后缀，避免 `splitext` 的问题
                        if rel_path.endswith(suffix):
                            rel_path_no_ext = rel_path[:-len(suffix)]
                        else:
                            rel_path_no_ext = os.path.splitext(rel_path)[0]

                        self.type_system.load_type_def(rel_path_no_ext.replace('\\', '/'), silent=True)
                    except ValueError as e:
                        print(f"警告: 无法解析 {file}。错误: {e}")

    def read_all(self) -> list[ConfigTable]:
        """读取 input_dir 中的所有 Excel 文件并返回 ConfigTable 对象列表。"""
        tables = []
        excel_files = [f for f in os.listdir(self.input_dir) if f.endswith('.xlsx') and not f.startswith('~')]
        
        for excel_file in excel_files:
            base_name = os.path.splitext(excel_file)[0]
            typedef_path = os.path.join(self.metadata_dir, f"{base_name}{self.typedef_suffix}")
            
            if not os.path.exists(typedef_path):
                print(f"提示: 找不到 '{excel_file}' 对应的 typedef 文件，已跳过。")
                continue
            
            with open(typedef_path, 'r', encoding='utf-8') as f:
                typedef_data = json.load(f)
            
            for imp in typedef_data.get("ImportTypes", []):
                self.type_system.load_type_def(imp)

            main_type_name = typedef_data["TargetType"]
            if not typedef_data.get("IsFlatTable"):
                if main_type_name not in self.type_system._loaded_types:
                    self.type_system._loaded_types[main_type_name] = typedef_data

            workbook = openpyxl.load_workbook(os.path.join(self.input_dir, excel_file), data_only=True)
            sheet = workbook.worksheets[0]
            
            # 优先使用 typedef 文件中定义的 Comment
            table_comment = typedef_data.get("Comment") or sheet.cell(row=1, column=1).value or f"由 {excel_file} 生成的配置"

            table = ConfigTable(
                excel_file_name=excel_file,
                base_name=base_name,
                is_flat_table=typedef_data.get("IsFlatTable", False),
                target_type_name=main_type_name,
                table_comment=table_comment
            )
            
            if table.is_flat_table:
                self._parse_flat_table(sheet, table)
            else:
                table.primary_key_fields = typedef_data.get("PrimaryKeyFields", [])
                self._parse_standard_table(sheet, typedef_data, table)
                
            tables.append(table)
            
        return tables

    def _parse_flat_table(self, sheet, table: ConfigTable):
        """解析平铺式表格。"""
        header = [c.value for c in sheet[1]]
        try:
            key_idx = header.index('Key')
            type_idx = header.index('Type')
            val_idx = header.index('Value')
            cmt_idx = header.index('Comment') if 'Comment' in header else -1
        except ValueError as e:
            raise ValueError(f"平铺表 '{table.excel_file_name}' 缺少必需的列: {e}。应包含 'Key', 'Type', 'Value', 'Comment'(可选)。")
            
        for row_data in sheet.iter_rows(min_row=2, values_only=True):
            if row_data[key_idx] is None:
                continue
            table.rows.append(ConfigRow(
                key=row_data[key_idx],
                type_syntax=row_data[type_idx],
                value=row_data[val_idx],
                comment=row_data[cmt_idx] if cmt_idx != -1 else ""
            ))
    
    def _parse_standard_table(self, sheet, typedef_data: dict, table: ConfigTable):
        """解析标准表格，使用新的 FieldSequence 结构。"""
        field_definitions = typedef_data.get("FieldSequence", [])
        
        # 将 typedef 中的定义转换为 ConfigRow 对象
        for field_def in field_definitions:
            field_name = field_def.get("Field")
            if not field_name:
                continue
            table.rows.append(ConfigRow(
                key=field_name,
                type_syntax=field_def.get("Type", "string"),
                comment=field_def.get("Comment", "")
            ))
        
        # 读取数据行
        headers = {c.value: i for i, c in enumerate(sheet[2]) if c.value}
        for row_data in sheet.iter_rows(min_row=3, values_only=True):
            if all(v is None for v in row_data):
                continue
            ordered_row = []
            for field in table.rows:
                col_idx = headers.get(field.key)
                if col_idx is None:
                    raise ValueError(f"'{table.excel_file_name}' 的 Excel 文件中缺少列 '{field.key}'。")
                ordered_row.append(row_data[col_idx])
            table.data_rows.append(ordered_row)
