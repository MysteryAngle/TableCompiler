# ==============================================================================
# TableCompiler
# Copyright (c) 2025, Alex Liao. All rights reserved.
#
# This file is part of the TableCompiler project, a tool designed to
# compile Excel configuration sheets into type-safe code and binary data
# for high-performance projects.
# ==============================================================================

# /config_generator/wizard.py
import click
import json
import datetime
import os
import openpyxl

# 由于此模块被动态加载，我们从父级导入必要的类
from .readers import TypeSystem

def _scan_existing_innertypes(type_system: TypeSystem, metadata_dir: str, inner_type_def_suffix: str):
    """
    预扫描并加载 metadata 目录下的所有 innertype 定义，以填充类型系统的选项。
    """
    for root, _, files in os.walk(metadata_dir):
        for file in files:
            if file.endswith(inner_type_def_suffix):
                try:
                    type_system.load_type_def(os.path.join(root, file), silent=True)
                except ValueError as e:
                    click.echo(click.style(f"警告: 无法解析 {file}。错误: {e}", fg='yellow'))

def _save_json(path: str, data: dict):
    """以美观的格式保存 JSON 文件。"""
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def _create_new_inner_type_interactive(type_system: TypeSystem, metadata_dir: str, inner_type_def_suffix: str) -> tuple[str, str]:
    """
    引导用户创建一个新的 innertype 定义，并将其保存到对应的 .innertypesdef.json 文件中。
    返回 (新类型名称, 相对导入路径)。
    """
    click.echo(click.style("\n--- 创建新的自定义类型 ---", fg='cyan'))
    name = click.prompt("为新类型输入唯一名称 (如 Item, Reward)", type=str)
    if not name:
        click.echo("类型名称不能为空。")
        return "", ""

    path = click.prompt("输入其生成代码中的目标路径 (如 Common/Item)", default=f"Common/{name}")
    comment = click.prompt(f"输入对此类型 '{name}' 的描述 (将作为类/枚举注释)", default=f"Represents a {name}.")
    rel_path = click.prompt("输入保存此定义的相对路径 (metadata/)", default=f"InnerTypes/{name}")
    abs_path = os.path.join(metadata_dir, f"{rel_path}{inner_type_def_suffix}")
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    
    is_enum = click.confirm(f"'{name}' 是一个枚举(Enum)吗?")
    new_def = { "TargetType": path, "Comment": comment }
    
    if is_enum:
        new_def["TargetTypeAsEnum"] = True
        m_str = click.prompt("输入枚举成员 (如 Prop=1,Equip=2)", type=str)
        try:
            new_def["EnumMembers"] = {k.strip(): int(v.strip()) for p in m_str.split(',') for k,v in [p.strip().split('=')]}
        except (ValueError, IndexError):
            raise ValueError("枚举成员格式错误，请使用 'Key1=Value1,Key2=Value2' 格式。")
    else: # 是一个类
        f_str = click.prompt("输入此类字段名, 逗号分隔 (如 ItemId,Count)", type=str)
        fields = [f.strip() for f in f_str.split(',') if f.strip()]
        new_def["FieldSequence"] = []
        for f in fields:
            field_def, _ = _define_field_interactive(f, type_system, metadata_dir, inner_type_def_suffix, f"Property {f} of {name}")
            if field_def:
                new_def["FieldSequence"].append(field_def)

    data = {"ImportTypes": [], "TypeDefines": { name: new_def } }
    
    if os.path.exists(abs_path) and not click.confirm(f"'{os.path.basename(abs_path)}' 已存在。要覆盖它吗?", default=False):
        click.echo("操作已取消。")
        return "", ""

    _save_json(abs_path, data)
    type_system._loaded_types[name] = new_def
    return name, rel_path

def _select_type_interactive(prompt_name: str, type_system: TypeSystem, metadata_dir: str, inner_type_def_suffix: str) -> tuple[str, list[str]]:
    """
    一个纯粹的、递归的函数，只负责通过交互确定一个类型字符串及其导入。
    """
    basic = ["int", "long", "string", "bool", "float"]
    coll = ["list", "array", "set"]
    custom = type_system.get_all_custom_type_names()
    choices = basic + coll + custom + ["[创建新的自定义类型]"]
    
    for i, choice in enumerate(choices):
        click.echo(f"  [{i+1}] {choice}")
    
    choice_idx = click.prompt(f"为 '{prompt_name}' 选择类型", type=int, default=1) - 1
    if not 0 <= choice_idx < len(choices):
        click.echo(click.style("无效选择，将默认为 string。", fg='yellow'))
        return "string", []

    selected = choices[choice_idx]

    if selected in basic:
        return selected, []
        
    if selected in coll:
        click.echo(f"--- 定义 {selected} '{prompt_name}' 的子项类型 ---")
        inner_type_str, inner_imps = _select_type_interactive(f"{prompt_name}.Item", type_system, metadata_dir, inner_type_def_suffix)
        final_type_str = f"{selected}({inner_type_str})"
        if click.confirm(f"此集合 '{final_type_str}' 需要从带分隔符的字符串解析吗?", default=False):
            delimiters = click.prompt("输入所有层级的分隔符, 空格分隔 (如 ~ #)", default="~ #")
            final_type_str += json.dumps(delimiters.split(" "))
        return final_type_str, inner_imps
        
    if selected == "[创建新的自定义类型]":
        name, path = _create_new_inner_type_interactive(type_system, metadata_dir, inner_type_def_suffix)
        if not name: return "string", [] # 用户取消了创建
        return name, [path]
        
    return selected, []

def _define_field_interactive(field_name: str, type_system: TypeSystem, metadata_dir: str, inner_type_def_suffix: str, default_comment: str = "") -> tuple[dict | None, list[str]]:
    """
    交互式地定义单个字段的所有属性（类型和注释）。
    返回 (字段定义字典, 新的导入列表)。
    """
    click.echo(f"\n--- 定义字段: {click.style(field_name, bold=True)} ---")
    
    final_type_str, new_imports = _select_type_interactive(field_name, type_system, metadata_dir, inner_type_def_suffix)
    if not final_type_str:
        return None, []
    
    comment = click.prompt(f"输入字段 '{field_name}' 的注释", default=default_comment)
    
    return {"Field": field_name, "Type": final_type_str, "Comment": comment}, new_imports

def _update_existing_typedef(data: dict, path: str, headers: list[str], header_comments: dict, metadata_dir: str, inner_type_def_suffix: str):
    """比对并更新一个已存在的标准表格 typedef 文件。"""
    type_system = TypeSystem()
    _scan_existing_innertypes(type_system, metadata_dir, inner_type_def_suffix)
    
    current_comment = data.get("Comment", "")
    new_comment = click.prompt("编辑主类注释", default=current_comment)
    data["Comment"] = new_comment

    field_sequence = data.get("FieldSequence", [])
    defined_fields_map = {d["Field"]: d for d in field_sequence}
    excel_fields = set(headers)
    
    new_fields = excel_fields - set(defined_fields_map.keys())
    removed_fields = set(defined_fields_map.keys()) - excel_fields
    
    if new_fields:
        click.echo(click.style(f"\n在Excel中找到新字段: {', '.join(new_fields)}", fg='yellow'))
        for f in new_fields:
            field_def, imps = _define_field_interactive(f, type_system, metadata_dir, inner_type_def_suffix, default_comment=header_comments.get(f, ""))
            if field_def:
                data["FieldSequence"].append(field_def)
            for imp in imps:
                if imp not in data["ImportTypes"]:
                    data["ImportTypes"].append(imp)

    if removed_fields:
        click.echo(click.style(f"在Excel中已移除的字段: {', '.join(removed_fields)}", fg='yellow'))
        if click.confirm("要从typedef中移除这些字段吗?", default=True):
            data["FieldSequence"] = [d for d in data["FieldSequence"] if d["Field"] not in removed_fields]

    if not new_fields and not removed_fields:
        click.echo("字段已同步。")

    if click.confirm("\n要手动编辑现有字段的类型或注释吗?", default=False):
        while True:
            click.echo("要修改哪个字段?")
            current_fields = data["FieldSequence"]
            for i, f_def in enumerate(current_fields):
                click.echo(f"  [{i+1}] {f_def['Field']} (Type: {f_def['Type']}, Comment: '{f_def.get('Comment', '')}')")
            idx = click.prompt("输入编号, 或输入0完成", type=int, default=0) - 1
            if idx < 0: break
            
            if 0 <= idx < len(current_fields):
                field_to_edit = current_fields[idx]
                updated_def, imps = _define_field_interactive(
                    field_to_edit['Field'], type_system, metadata_dir, inner_type_def_suffix, default_comment=field_to_edit.get('Comment', '')
                )
                if updated_def:
                    data["FieldSequence"][idx] = updated_def
                for imp in imps:
                    if imp not in data["ImportTypes"]:
                        data["ImportTypes"].append(imp)

    _save_json(path, data)
    click.echo(click.style(f"\n完成更新 '{os.path.basename(path)}'", fg='green'))

def _create_new_typedef(typedef_path: str, base_name: str, input_dir: str, metadata_dir: str, inner_type_def_suffix: str):
    """引导用户创建一个全新的 typedef 文件。"""
    click.echo("--- 定义主类型 ---")
    target_type_name = click.prompt("输入主类名", default=base_name)
    
    data = {
        "ExcelFileName": f"{base_name}.xlsx",
        "Version": datetime.datetime.now().strftime("%Y%m%d_%H%M%S"),
        "TargetType": target_type_name,
        "Comment": click.prompt(f"输入对此主类 '{target_type_name}' 的描述", default=f"Represents a {target_type_name} configuration."),
        "ImportTypes": []
    }
    data["IsFlatTable"] = click.confirm("这是一个平铺式表格 (Key-Value) 吗?", default=False)

    if not data["IsFlatTable"]:
        excel_path = os.path.join(input_dir, data["ExcelFileName"])
        try:
            sheet = openpyxl.load_workbook(excel_path, data_only=True).worksheets[0]
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到对应的Excel文件: {excel_path}")
            
        headers = [c.value for c in sheet[2] if c.value]
        header_comments = {sheet.cell(row=2, column=i+1).value: sheet.cell(row=1, column=i+1).value for i, h in enumerate(sheet[2]) if h.value}
        
        pk_str = click.prompt(f"从 [{', '.join(headers)}] 中输入主键字段, 逗号分隔", default=headers[0] if headers else "Id")
        data["PrimaryKeyFields"] = [k.strip() for k in pk_str.split(',')]
        data["FieldSequence"] = []
        
        type_system = TypeSystem()
        _scan_existing_innertypes(type_system, metadata_dir, inner_type_def_suffix)
        
        for h in headers:
            field_def, imps = _define_field_interactive(h, type_system, metadata_dir, inner_type_def_suffix, default_comment=header_comments.get(h, ""))
            if field_def:
                data["FieldSequence"].append(field_def)
            for imp in imps:
                if imp not in data["ImportTypes"]:
                    data["ImportTypes"].append(imp)

    _save_json(typedef_path, data)
    click.echo(click.style(f"\n成功创建 '{os.path.basename(typedef_path)}'", fg='green'))

@click.command(name="typedef")
@click.pass_context
def typedef_command(ctx):
    """用于创建或更新 .typedef.json 文件的交互式向导。"""
    input_dir = ctx.obj['INPUT_DIR']
    metadata_dir = ctx.obj['METADATA_DIR']
    typedef_suffix = ctx.obj['TYPE_DEF_SUFFIX']
    inner_type_def_suffix = ctx.obj['INNER_TYPE_DEF_SUFFIX']
    
    click.echo(click.style("===== Typedef 向导 =====", bold=True))
    excel_files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx') and not f.startswith('~')]
    if not excel_files:
        click.echo(click.style(f"在 '{input_dir}' 目录中未找到 .xlsx 文件。", fg='yellow'))
        return

    click.echo("请选择一个Excel文件来管理其typedef:")
    for i, f in enumerate(excel_files):
        click.echo(f"  [{i+1}] {f}")

    choice_idx = click.prompt("输入您的选择编号", type=int, default=1) - 1
    if not 0 <= choice_idx < len(excel_files):
        click.echo(click.style("无效选择。", fg='red'))
        return

    excel_file = excel_files[choice_idx]
    excel_path = os.path.join(input_dir, excel_file)
    base_name = os.path.splitext(excel_file)[0]
    typedef_path = os.path.join(metadata_dir, f"{base_name}{typedef_suffix}")

    if not os.path.exists(typedef_path):
        click.echo(f"未找到 '{excel_file}' 的typedef。启动创建向导...")
        _create_new_typedef(typedef_path, base_name, input_dir, metadata_dir, inner_type_def_suffix)
    else:
        click.echo(f"找到 '{excel_file}' 的现有typedef。检查更新...")
        with open(typedef_path, 'r', encoding='utf-8') as f:
            typedef_data = json.load(f)
        
        if typedef_data.get("IsFlatTable"):
            click.echo(click.style("这是一个平铺式表格的Typedef, 其结构定义在Excel内部, 无需更新。", fg='blue'))
        else:
            sheet = openpyxl.load_workbook(excel_path, data_only=True).worksheets[0]
            excel_headers = [c.value for c in sheet[2] if c.value]
            header_comments = {sheet.cell(row=2, column=i+1).value: sheet.cell(row=1, column=i+1).value for i, h in enumerate(sheet[2]) if h.value}
            _update_existing_typedef(typedef_data, typedef_path, excel_headers, header_comments, metadata_dir, inner_type_def_suffix)
